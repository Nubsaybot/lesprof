from openpyxl import load_workbook
import undetected_chromedriver as uc
import time as tm
from bs4 import BeautifulSoup
import re
import json

def suppress_exception_in_del(uc):
    old_del = uc.Chrome.__del__

    def new_del(self) -> None:
        try:
            old_del(self)
        except:
            pass
    
    setattr(uc.Chrome, '__del__', new_del)

suppress_exception_in_del(uc)

uc.TARGET_VERSION = 124

driver = uc.Chrome(headless=False, use_subprocess=True, )
driver.implicitly_wait(10)

mainData_book = load_workbook('lesprof.xlsx')
sheet = mainData_book.active

print(sheet.max_row, sheet.max_column)
for column in range(1, sheet.max_column+1):
    for row in range(1, sheet.max_row+1):
        # if sheet.cell(row,column).value != None:
        if sheet.cell(row,column).hyperlink:
            url = sheet.cell(row,column).hyperlink.target
            # url = str(sheet.cell(row,column).value)
            driver.get(url)
            tm.sleep(4)
            page_source = str(driver.page_source)
            page_source=re.sub(r'<!.*?->','', page_source)
            # print(page_source)
            # f = open('xyz.txt','w', encoding="utf-8")  # открытие в режиме записи
            # f.write(page_source)
            # f.close()  # закрытие файла
            soup = BeautifulSoup(driver.page_source, 'html.parser')

            try:
                data = soup.select_one(f'div[id^=state-webPrice]')['data-state']
                json_data = json.loads(data)
                if 'originalPrice' in json_data:
                    base_price = int(json_data['originalPrice'].encode('ascii','ignore'))
                else:
                    base_price = 0
                if 'price' in json_data:
                    discount_price = int(json_data['price'].encode('ascii','ignore'))
                else:
                    discount_price = 0
                if 'cardPrice' in json_data:
                    ozon_card_price = int(json_data['cardPrice'].encode('ascii','ignore'))
                else:
                    ozon_card_price = 0
                if 'pricePerUnit' in json_data:
                    pricePerUnit = int(json_data['pricePerUnit'].encode('ascii','ignore'))
                else:
                    pricePerUnit = 0
                sheet.cell(row, column+1).value = discount_price
                # sheet.cell(row, column+2).value = base_price
    #            sheet.cell(row, column+3).value = ozon_card_price
    #            sheet.cell(row, column+4).value = pricePerUnit
            except:
                discount_price = 0
                base_price = 0
                ozon_card_price = 0
                pricePerUnit = 0

            # print("Столбец=" + column + "    Строка=" + column + "    Базовая=" + str(base_price) + "    Со скидкой=" + str(discount_price) + "    По карте=" + str(ozon_card_price) + "    За штуку=" + str(pricePerUnit))
            print("Столбец="+str(column)+"  Строка="+str(row))

mainData_book.save('lesprof.xlsx')
